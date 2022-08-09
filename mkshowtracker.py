#! /usr/bin/python3

#Programmed with python 3.8.9
#Ananlyzes a progress report and makes a tracker for Ronnie Perez

import os
import re
import sys
import json
import airtable
import requests
import openpyxl

apikey    = "keynFUEFq7QZPS34A"
baseid    = "appJAa0wWijCR3dpm"
headers   = {"Authorization" : "Bearer " + apikey, "Content-Type" : "application/json"}

atrecords = { "records": [] }
channeldict = {}

#get the AirTable Base
def getAtTableName():

    tablename = input("  Give me your Air Table Tracker Name: ")
    return tablename

#get the name of the xlf (Progress Report)
def getXLF():

    print("")
    while True:
        xlf = input("  Give me your progress report (.xlsx extension): ")

        m = re.search("^(.+)\.xlsx$",xlf)
        if m and os.path.exists(xlf):
            prefix = m.group(1)
            break
    return xlf,prefix

#select the worksheet from the Progress Report
def selectWS(wb,xlf):

    print("")
    while True:

        print("  Work Sheets in", xlf)
        print("  ==============")
        for sheet in wb.sheetnames:
            print(" ",sheet)
        sheet = input("\n  Select Your Work Sheet : ")
        if sheet in wb.sheetnames:
            break
    print("")
    return sheet

def readXLF(ws):

    r         = 1
    showcol   = 1
    epcol     = 2
    nonecount = 0
    showdict  = {}
    sheetdata = ws['A1':'B5000']

    for n in range(0,len(sheetdata)):
        show = str(sheetdata[n][0].value)
        snum = str(sheetdata[n][1].value)
        #print(show,snum)

        if show != 'None':
            season = show + ':' + snum.zfill(2)
            #incriment the value by 1, so we cna count number of episodes
            if season in showdict:
                n = showdict[season]
                n += 1
                showdict[season] = n
                #if this is the first time we enountered show:## set the value (show counter to 1)
            else:
                if show != 'Show Title':
                    showdict[season] = 1
        r += 1

    return showdict

#def postData(atrecords, aturl):
#
#    for i in range(0, len(atrecords["records"]) ):
#        if (i + 1) % 10 == 0:





####Main Program####
xlf,prefix = getXLF()

#Open the xlf / Progress Report
try:
    wb = openpyxl.load_workbook(filename=xlf, read_only=True)
except:
    print("  ~~Could not open", xlf,"\n")
    sys.exit(1)

sheet = selectWS(wb,xlf)
ws = wb[sheet]

channeldict = readXLF(ws)
wb.close()

tablename = getAtTableName()
aturl = "https://api.airtable.com/v0/" + baseid + "/" + tablename
#print(aturl)

i    = 1
oldi = i
print("")
#print("Show:Season:# Episodes:Notes:Merge Captions / MXF:Jarvis:Upload XL -> S3:Caption -> S3:Caption -> Box Archive:Cleared V1 In Veritone:Status")
for s in sorted(channeldict.keys()):
    if s != "Show Title::Season Number:":
        numeps = channeldict[s]
        showParts = s.split(':')
        #season = showParts[0] + ":" + showParts[1] + ":" + str(numeps) + "\n"
        season = showParts[0] + ":" + showParts[1] + ":" + str(numeps)
        #print(season)

        rec = { 'fields': {'Show' : showParts[0], 'Season' : showParts[1], '# Episodes' : str(numeps) } }
        atrecords["records"].append(rec)

        if i % 10 == 0:
            print("  Pushing Records",oldi,"->",i,"to Airtable",end=" ",flush=True)
            response = requests.request("POST", aturl, headers=headers, data=json.dumps(atrecords))
            if not response:
                print(": ERROR")
            else:
                print("\n")
            atrecords = { "records": [] }
            oldi = i + 1
        i += 1

if atrecords != { "records": [] }:
    print("  Pushing Records",oldi,"->",i,"to Airtable",end=" ",flush=True)
    response = requests.request("POST", aturl, headers=headers, data=json.dumps(atrecords))
    if not response:
        print(": ERROR")
    else:
        print("\n")


"""
#print(atrecords)
print("atrcords               :",type(atrecords))
print("=======")
print("atrecords[\"records\"]   :",type(atrecords["records"]))
print("=======")
print("fields                 :",type(rec))
"""
